import io
import openpyxl
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from accounts.models import Branch
from academics.models import Level, SchoolClass, AcademicYear
from students.models import Student


def _filter_branch(qs, user):
    if not user.is_admin and user.branch_id:
        return qs.filter(branch_id=user.branch_id)
    return qs


def _generate_student_id():
    base = f"STU-{timezone.now().strftime('%Y%m%d')}"
    counter = 1
    candidate = f"{base}-{counter:04d}"
    while Student.objects.filter(student_id=candidate).exists():
        counter += 1
        candidate = f"{base}-{counter:04d}"
    return candidate


@login_required
def student_list(request):
    qs = _filter_branch(Student.objects.select_related("branch", "level", "school_class"), request.user)

    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(student_id__icontains=q) |
            Q(parent_phone__icontains=q)
        )

    branch_id = request.GET.get("branch")
    if branch_id:
        qs = qs.filter(branch_id=branch_id)

    class_id = request.GET.get("class")
    if class_id:
        qs = qs.filter(school_class_id=class_id)

    students = qs.all()
    branches = Branch.objects.all()
    classes = SchoolClass.objects.all()

    return render(request, "students/student_list.html", {
        "students": students,
        "branches": branches,
        "classes": classes,
        "q": q,
    })


@login_required
def student_detail(request, pk):
    student = get_object_or_404(Student.objects.select_related("branch", "level", "school_class", "academic_year"), pk=pk)
    payments = student.payments.select_related("fee_type", "accountant").all()

    # Calculate financial summary
    from finance.models import FeeAssignment
    from payments.models import Payment
    active_year = AcademicYear.objects.filter(is_active=True).first()
    assignments = FeeAssignment.objects.filter(
        Q(scope="school") |
        Q(scope="branch", branch=student.branch) |
        Q(scope="level", level=student.level) |
        Q(scope="class", school_class=student.school_class)
    ).filter(academic_year=active_year).select_related("fee_type") if active_year else []

    fee_summary = []
    for fa in assignments:
        paid = Payment.objects.filter(
            student=student, fee_type=fa.fee_type,
            status=Payment.Status.SUCCESSFUL
        ).aggregate(total=Sum("amount"))["total"] or 0
        remaining = fa.amount - paid
        status = "Fully Paid" if remaining <= 0 else ("Partial" if paid > 0 else "Unpaid")
        fee_summary.append({
            "fee_type": fa.fee_type.name,
            "required": fa.amount,
            "paid": paid,
            "remaining": remaining,
            "status": status,
        })

    return render(request, "students/student_detail.html", {
        "student": student,
        "payments": payments,
        "fee_summary": fee_summary,
    })


@login_required
def student_create(request):
    branches = Branch.objects.all()
    levels = Level.objects.all()
    classes = SchoolClass.objects.select_related("level", "branch").all()
    years = AcademicYear.objects.all()

    if request.method == "POST":
        student_id = request.POST.get("student_id", "").strip() or _generate_student_id()
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        gender = request.POST.get("gender")
        date_of_birth = request.POST.get("date_of_birth") or None
        parent_name = request.POST.get("parent_name", "").strip()
        parent_phone = request.POST.get("parent_phone", "").strip()
        branch_id = request.POST.get("branch")
        level_id = request.POST.get("level")
        class_id = request.POST.get("school_class")
        year_id = request.POST.get("academic_year")

        # Auto-set branch for non-admin
        if not request.user.is_admin and request.user.branch_id:
            branch_id = request.user.branch_id

        branch = Branch.objects.filter(pk=branch_id).first()
        level = Level.objects.filter(pk=level_id).first()
        school_class = SchoolClass.objects.filter(pk=class_id).first()
        year = AcademicYear.objects.filter(pk=year_id).first()

        if not all([first_name, last_name, branch, level, school_class]):
            messages.error(request, "Name, branch, level, and class are required.")
            return render(request, "students/student_form.html", {
                "branches": branches, "levels": levels, "classes": classes, "years": years, "title": "Add Student", "generated_student_id": _generate_student_id()
            })

        if Student.objects.filter(student_id=student_id).exists():
            messages.error(request, f"Student ID '{student_id}' already exists.")
            return render(request, "students/student_form.html", {
                "branches": branches, "levels": levels, "classes": classes, "years": years, "title": "Add Student", "generated_student_id": _generate_student_id()
            })

        Student.objects.create(
            student_id=student_id, first_name=first_name, last_name=last_name,
            gender=gender, date_of_birth=date_of_birth,
            parent_name=parent_name, parent_phone=parent_phone,
            branch=branch, level=level, school_class=school_class,
            academic_year=year,
        )
        messages.success(request, f"Student '{first_name} {last_name}' registered.")
        return redirect("students:student_list")

    return render(request, "students/student_form.html", {
        "branches": branches, "levels": levels, "classes": classes, "years": years, "title": "Add Student", "generated_student_id": _generate_student_id()
    })


@login_required
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    branches = Branch.objects.all()
    levels = Level.objects.all()
    classes = SchoolClass.objects.select_related("level", "branch").all()
    years = AcademicYear.objects.all()

    if request.method == "POST":
        student.student_id = request.POST.get("student_id", student.student_id).strip()
        student.first_name = request.POST.get("first_name", student.first_name).strip()
        student.last_name = request.POST.get("last_name", student.last_name).strip()
        student.gender = request.POST.get("gender", student.gender)
        student.date_of_birth = request.POST.get("date_of_birth") or None
        student.parent_name = request.POST.get("parent_name", "").strip()
        student.parent_phone = request.POST.get("parent_phone", "").strip()
        student.branch = Branch.objects.filter(pk=request.POST.get("branch")).first() or student.branch
        student.level = Level.objects.filter(pk=request.POST.get("level")).first() or student.level
        student.school_class = SchoolClass.objects.filter(pk=request.POST.get("school_class")).first() or student.school_class
        student.academic_year = AcademicYear.objects.filter(pk=request.POST.get("academic_year")).first()
        student.status = request.POST.get("status", student.status)
        student.save()
        messages.success(request, "Student updated.")
        return redirect("students:student_detail", student.pk)

    return render(request, "students/student_form.html", {
        "student": student, "branches": branches, "levels": levels, "classes": classes, "years": years, "title": "Edit Student"
    })


@login_required
@login_required
def student_import(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
        except Exception as e:
            messages.error(request, f"Error reading file: {e}")
            return redirect("students:student_import")

        results = {"success": [], "errors": []}
        
        # Use atomic transaction to ensure all-or-nothing import
        try:
            with transaction.atomic():
                for i, row in enumerate(rows, start=2):
                    if not any(row):
                        continue
                    try:
                        student_id = str(row[0] or "").strip()
                        first_name = str(row[1] or "").strip()
                        last_name = str(row[2] or "").strip()
                        gender = str(row[3] or "").strip().upper()[:1]
                        dob = row[4] or None
                        parent_name = str(row[5] or "").strip()
                        parent_phone = str(row[6] or "").strip()
                        branch_code = str(row[7] or "").strip()
                        level_code = str(row[8] or "").strip()
                        class_name = str(row[9] or "").strip()

                        if not student_id or not first_name or not last_name:
                            results["errors"].append(f"Row {i}: Missing required fields")
                            continue
                        if Student.objects.filter(student_id=student_id).exists():
                            results["errors"].append(f"Row {i}: Student ID '{student_id}' already exists")
                            continue

                        branch = Branch.objects.filter(code=branch_code).first()
                        if not branch:
                            results["errors"].append(f"Row {i}: Branch '{branch_code}' not found")
                            continue

                        level = Level.objects.filter(code=level_code).first()
                        if not level:
                            results["errors"].append(f"Row {i}: Level '{level_code}' not found")
                            continue

                        school_class = SchoolClass.objects.filter(name=class_name, level=level).first()
                        if not school_class:
                            results["errors"].append(f"Row {i}: Class '{class_name}' not found for level '{level_code}'")
                            continue

                        Student.objects.create(
                            student_id=student_id, first_name=first_name, last_name=last_name,
                            gender=gender, date_of_birth=dob,
                            parent_name=parent_name, parent_phone=parent_phone,
                            branch=branch, level=level, school_class=school_class,
                        )
                        results["success"].append(f"Row {i}: {first_name} {last_name}")
                    except Exception as e:
                        results["errors"].append(f"Row {i}: {e}")
        except Exception as e:
            messages.error(request, f"Import failed: {e}. No students were imported.")
            return redirect("students:student_import")

        messages.success(request, f"Imported {len(results['success'])} students.")
        if results["errors"]:
            messages.warning(request, f"{len(results['errors'])} errors: " + "; ".join(results["errors"][:5]))

        return render(request, "students/import_result.html", {"results": results})

    branches = Branch.objects.all()
    levels = Level.objects.all()
    classes = SchoolClass.objects.all()
    return render(request, "students/student_import.html", {"branches": branches, "levels": levels, "classes": classes})


@login_required
def student_export(request):
    qs = _filter_branch(Student.objects.select_related("branch", "level", "school_class"), request.user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["Student ID", "First Name", "Last Name", "Gender", "Date of Birth", "Parent Name", "Parent Phone", "Branch", "Level", "Class", "Status"])

    for s in qs:
        ws.append([
            s.student_id, s.first_name, s.last_name, s.gender,
            s.date_of_birth.isoformat() if s.date_of_birth else "",
            s.parent_name, s.parent_phone, s.branch.name, s.level.name,
            s.school_class.name, s.status
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=students.xlsx"
    wb.save(response)
    return response


@login_required
def student_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["Student ID", "First Name", "Last Name", "Gender (M/F)", "Date of Birth (YYYY-MM-DD)", "Parent Name", "Parent Phone", "Branch Code", "Level Code", "Class Name"])
    ws.append(["STU001", "John", "Doe", "M", "2015-03-15", "Jane Doe", "0781234567", "MAIN", "PRIM", "P1"])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=student_template.xlsx"
    wb.save(response)
    return response


@login_required
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if not request.user.is_admin and request.user.branch_id != student.branch_id:
        messages.error(request, "Access denied.")
        return redirect("students:student_list")
    if request.method == "POST":
        name = student.full_name
        student.delete()
        messages.success(request, f"Student '{name}' deleted.")
        return redirect("students:student_list")
    return render(request, "students/student_confirm_delete.html", {"student": student})

