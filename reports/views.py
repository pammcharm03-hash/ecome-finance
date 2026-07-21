import openpyxl
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum
from django.utils import timezone
from datetime import timedelta

from accounts.models import Branch
from students.models import Student
from payments.models import Payment
from finance.models import FeeType


def _filter_branch(qs, user):
    if not user.is_admin and user.branch_id:
        return qs.filter(branch_id=user.branch_id)
    return qs


@login_required
def home(request):
    if not request.user.is_admin and not request.user.is_accountant:
        messages.error(request, "Access denied.")
        return redirect("dashboard:home")
    return render(request, "reports/home.html")


@login_required
def export_students(request):
    if not request.user.is_admin and not request.user.is_registrar:
        messages.error(request, "Access denied.")
        return redirect("dashboard:home")

    qs = _filter_branch(Student.objects.select_related("branch", "level", "school_class"), request.user)

    branch_id = request.GET.get("branch")
    if branch_id and request.user.is_admin:
        qs = qs.filter(branch_id=branch_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = ["Student ID", "First Name", "Last Name", "Gender", "Date of Birth", "Parent Name", "Parent Phone", "Branch", "Level", "Class", "Status"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

    for s in qs:
        ws.append([
            s.student_id, s.first_name, s.last_name, s.gender,
            s.date_of_birth.isoformat() if s.date_of_birth else "",
            s.parent_name, s.parent_phone, s.branch.name, s.level.name,
            s.school_class.name, s.status
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=students_report.xlsx"
    wb.save(response)
    return response


@login_required
def export_payments(request):
    if not request.user.is_admin and not request.user.is_accountant:
        messages.error(request, "Access denied.")
        return redirect("dashboard:home")

    qs = _filter_branch(
        Payment.objects.select_related("student", "fee_type", "branch", "accountant").filter(status=Payment.Status.SUCCESSFUL),
        request.user
    )

    period = request.GET.get("period", "all")
    today = timezone.now().date()
    if period == "daily":
        qs = qs.filter(created_at__date=today)
    elif period == "weekly":
        qs = qs.filter(created_at__date__gte=today - timedelta(days=7))
    elif period == "monthly":
        qs = qs.filter(created_at__date__gte=today.replace(day=1))
    elif period == "yearly":
        qs = qs.filter(created_at__year=today.year)

    branch_id = request.GET.get("branch")
    if branch_id and request.user.is_admin:
        qs = qs.filter(branch_id=branch_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"
    headers = ["Receipt No", "Student ID", "Student Name", "Fee Type", "Amount", "Phone", "Accountant", "Branch", "Date", "HDEV Ref"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

    total = 0
    for p in qs:
        ws.append([
            p.receipt_number, p.student.student_id, p.student.full_name,
            p.fee_type.name, float(p.amount), p.parent_phone,
            p.accountant.get_full_name() if p.accountant else "-",
            p.branch.name, p.created_at.strftime("%Y-%m-%d %H:%M"),
            p.paypack_ref or p.transaction_ref or "-",
        ])
        total += float(p.amount)

    ws.append([])
    ws.append(["", "", "", "TOTAL", total, "", "", "", "", ""])
    ws.cell(row=ws.max_row, column=5).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=ws.max_row, column=4).font = openpyxl.styles.Font(bold=True)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=payments_report_{period}.xlsx"
    wb.save(response)
    return response
